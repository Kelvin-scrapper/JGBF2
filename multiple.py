#!/usr/bin/env python3
"""
JGBF Excel Parser - Final Production Version (Custom Target Format)
Reads extracted Excel files and generates a single, consolidated JGBF_DATA
output file with the exact column structure specified by the user's template.
"""

import os
import sys
from pathlib import Path
import logging
import re
from typing import Dict, List, Tuple, Optional
from datetime import datetime
import pandas as pd

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font
except ImportError:
    print("Missing required packages. Please install with:")
    print("pip install openpyxl pandas")
    sys.exit(1)

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)


class JGBFParser:
    """Parses Excel files and generates a consolidated JGBF_DATA output."""

    def __init__(self, output_folder: str = "parsed_output"):
        self.output_folder = Path(output_folder)
        self.output_folder.mkdir(exist_ok=True)
        self.instrument_mapping = {"JGB(10-year) Futures": "JGB10YEARFUTURES", "mini-10-year JGB Futures": "MINI10YEARJGBFUTURESCASHSETTLED", "mini-20-year JGB Futures": "MINI20YEARJGBFUTURES", "3-Month TONA Futures": "3MONTHTONAFUTURES"}
        self.main_summary_categories = {"自己取引計": "PROPRIETARY", "委託取引計": "BROKERAGE", "自己委託合計": "TOTAL"}
        self.brokerage_categories = {"法人計": "INSTITUTIONS", "個人計": "INDIVIDUALS", "海外投資家計": "FOREIGNERS", "証券会社": "SECURITIES_COS"}
        self.subcategories = {"売り": "SALES", "買い": "PURCHASES"}
        self.month_map = {"Jan": 1, "Feb": 2, "Mar": 3, "Apr": 4, "May": 5, "Jun": 6, "Jul": 7, "Aug": 8, "Sep": 9, "Oct": 10, "Nov": 11, "Dec": 12}
        logger.info(f"[INFO] JGBF Parser initialized. Output will be saved to: {self.output_folder}")

    def get_template_columns(self) -> List[Dict]:
        """
        Return the exact column structure from the user's target CSV format.
        """
        full_template = []
        instruments = {
            "JGB10YEARFUTURES": "JGB(10-year) Futures",
            "MINI10YEARJGBFUTURESCASHSETTLED": "mini-10-year JGB Futures（Cash-Settled)",
            "MINI20YEARJGBFUTURES": "mini-20-year JGB Futures",
            "3MONTHTONAFUTURES": "3-Month TONA Futures"
        }

        main_cats = {"PROPRIETARY": "Proprietary", "BROKERAGE": "Brokerage", "TOTAL": "Total"}
        for instr_code, instr_name in instruments.items():
            for cat_code, cat_name in main_cats.items():
                for subcat in ["SALES", "PURCHASES"]:
                    for metric in ["VALUE", "BALANCE"]:
                        code = f"JGBF.TOTAL_PROPRIETARY_BROKERAGE.{instr_code}.TRADINGVALUE.{cat_code}.{subcat}.{metric}.W"
                        desc = f"Trading by Type of Investors, {instr_name}, Total, Proprietary ＆ Brokerage, Trading Value, {cat_name}, {subcat.title()}, {metric.title()}"
                        full_template.append({'code': code, 'description': desc})

        breakdown_cats = {"INSTITUTIONS": "Institutions", "INDIVIDUALS": "Individuals", "FOREIGNERS": "Foreigners", "SECURITIES_COS": "Securities Cos"}
        for instr_code, instr_name in instruments.items():
            for cat_code, cat_name in breakdown_cats.items():
                for subcat in ["SALES", "PURCHASES"]:
                    for metric in ["VALUE", "BALANCE"]:
                        if instr_code == 'JGB10YEARFUTURES':
                            code = f"JGBF.TRADINGBYTYPEOFINVESTORS.{instr_code}.BREAKDOWNOFBROKERAGE.TRADINGVALUE.{cat_code}.{subcat}.{metric}.W"
                        else:
                            code = f"JGBF.TOTAL_PROPRIETARY_BROKERAGE.{instr_code}.BREAKDOWNOFBROKERAGE.TRADINGVALUE.{cat_code}.{subcat}.{metric}.W"
                        
                        desc = f"Trading by Type of Investors, {instr_name}, Breakdown of Brokerage, Trading Value, {cat_name}, {subcat.title()}, {metric.title()}"
                        full_template.append({'code': code, 'description': desc})
                        
        return full_template


    def extract_instrument_from_subtitle(self, subtitle: str) -> Optional[str]:
        subtitle_clean = subtitle.replace("（", "(").replace("）", ")").replace("、", ",")
        if "mini-20-year" in subtitle_clean.lower() or "超長期国債先物" in subtitle_clean: return "MINI20YEARJGBFUTURES"
        if "3-Month TONA" in subtitle_clean or "TONA" in subtitle_clean: return "3MONTHTONAFUTURES"
        if "JGB(10-year)" in subtitle_clean or "長期国債先物" in subtitle_clean:
            return "MINI10YEARJGBFUTURESCASHSETTLED" if "mini" in subtitle_clean.lower() or "ミニ" in subtitle_clean else "JGB10YEARFUTURES"
        logger.warning(f"[WARN] Could not map subtitle to instrument: {subtitle}"); return None

    def extract_date_from_filename(self, filename: str) -> str:
        match1 = re.search(r"_(\d{4})_Week\d_(\d{1,2})-(\d{1,2})", filename, re.IGNORECASE)
        if match1:
            try:
                year_str, month_str, day_str = match1.groups(); date_obj = datetime(int(year_str), int(month_str), int(day_str)); iso_year, iso_week, _ = date_obj.isocalendar(); return f"{iso_year}-{iso_week:02d}"
            except Exception: pass
        match2 = re.search(r"([A-Za-z]{3})_(\d{4})_Week\d_(\d{1,2})-\d{1,2}", filename, re.IGNORECASE)
        if match2:
            try:
                month_name, year_str, day_str = match2.groups(); month = self.month_map.get(month_name.title())
                if month: date_obj = datetime(int(year_str), month, int(day_str)); iso_year, iso_week, _ = date_obj.isocalendar(); return f"{iso_year}-{iso_week:02d}"
            except Exception: pass
        match3 = re.search(r"_([A-Za-z]{3})\.?_(\d{1,2})_(\d{4})_-_", filename, re.IGNORECASE)
        if match3:
            try:
                month_name, day_str, year_str = match3.groups(); month = self.month_map.get(month_name.title())
                if month: date_obj = datetime(int(year_str), month, int(day_str)); iso_year, iso_week, _ = date_obj.isocalendar(); return f"{iso_year}-{iso_week:02d}"
            except Exception: pass
        logger.warning(f"Could not extract a known date format from filename: {filename}"); return "UNKNOWN_DATE"

    def handle_negative_values(self, value: any) -> str:
        if value is None or value == "-" or value == "": return ""
        value_str = str(value).strip()
        return "-" + value_str[1:] if value_str.startswith("▲") else value_str

    def read_excel_sheet(self, file_path: Path, sheet_name: str) -> Optional[Dict]:
        try:
            wb = load_workbook(file_path, data_only=True, read_only=True)
            if sheet_name not in wb.sheetnames: return None
            ws = wb[sheet_name]
            subtitle = (ws['A2'].value or "").replace("Subtitle: ", "")
            data_rows = [row for row in ws.iter_rows(min_row=6, values_only=True) if row and row[0]]
            return {'subtitle': subtitle, 'data_rows': data_rows}
        except Exception as e:
            logger.error(f"Error reading sheet {sheet_name} from {file_path}: {e}")
            return None

    def _parse_table(self, sheet_data: Dict, instrument_code: str, date_code: str, table_type: str) -> List[Dict]:
        results = []
        is_main = table_type == "main_summary"
        category_map = self.main_summary_categories if is_main else self.brokerage_categories
        
        for row in sheet_data['data_rows']:
            if len(row) < 8: continue
            cat_full, subcat_raw, val, bal = str(row[0] or ""), str(row[1] or ""), row[5], row[7]
            if not cat_full or not subcat_raw or "合計" in subcat_raw: continue
            
            category = next((en for jp, en in category_map.items() if jp in cat_full), None)
            subcat = next((en for jp, en in self.subcategories.items() if jp in subcat_raw), None)
            if not category or not subcat: continue
            
            for metric, data_val in [("VALUE", val), ("BALANCE", bal)]:
                processed_value = self.handle_negative_values(data_val)
                if processed_value:
                    code = ""
                    if is_main:
                        code = f"JGBF.TOTAL_PROPRIETARY_BROKERAGE.{instrument_code}.TRADINGVALUE.{category}.{subcat}.{metric}.W"
                    else: # Brokerage Breakdown logic
                        if instrument_code == 'JGB10YEARFUTURES':
                            code = f"JGBF.TRADINGBYTYPEOFINVESTORS.{instrument_code}.BREAKDOWNOFBROKERAGE.TRADINGVALUE.{category}.{subcat}.{metric}.W"
                        else:
                            code = f"JGBF.TOTAL_PROPRIETARY_BROKERAGE.{instrument_code}.BREAKDOWNOFBROKERAGE.TRADINGVALUE.{category}.{subcat}.{metric}.W"
                    
                    if code:
                        results.append({'code': code, 'date': date_code, 'value': processed_value})
        return results

    def process_single_file(self, file_path: Path) -> List[Dict]:
        logger.info(f"-> Processing file: {file_path.name}")
        date_code = self.extract_date_from_filename(file_path.stem)
        if date_code == "UNKNOWN_DATE":
            logger.error(f"   Skipping file due to unknown date format: {file_path.name}")
            return []
        
        file_results = []
        try:
            wb = load_workbook(file_path, data_only=True, read_only=True)
            for sheet_name in wb.sheetnames:
                table_type = "main_summary" if "Table1_Main_Summary" in sheet_name else "brokerage_breakdown" if "Table2_Brokerage_Bre" in sheet_name else None
                if not table_type: continue
                
                sheet_data = self.read_excel_sheet(file_path, sheet_name)
                if not sheet_data: continue
                
                instrument_code = self.extract_instrument_from_subtitle(sheet_data['subtitle'])
                if not instrument_code: continue
                
                parsed_data = self._parse_table(sheet_data, instrument_code, date_code, table_type)
                file_results.extend(parsed_data)
        except Exception as e: logger.error(f"   Critical error processing file {file_path.name}: {e}")
        return file_results

    def generate_output_file(self, all_data: List[Dict], output_filename: str):
        if not all_data:
            logger.warning("No data was extracted to generate an output file.")
            return
        
        data_pivot = {}
        for item in all_data:
            data_pivot.setdefault(item['code'], {})[item['date']] = item['value']
        
        all_dates = sorted([d for d in {item['date'] for item in all_data} if d != "UNKNOWN_DATE"])
        if not all_dates:
            logger.error("No valid dates found in data. Cannot generate output."); return

        template_columns = self.get_template_columns()
        
        wb = Workbook()
        ws = wb.active
        ws.title = "JGBF_DATA"
        header_font = Font(bold=True)
        
        ws['A1'], ws['A1'].font = "Date", header_font
        ws['A2'], ws['A2'].font = "Description", header_font
        
        for col_idx, col_def in enumerate(template_columns, start=2):
            ws.cell(row=1, column=col_idx, value=col_def['code']).font = header_font
            ws.cell(row=2, column=col_idx, value=col_def.get('description', '')).font = header_font
            
        for row_idx, date in enumerate(all_dates, start=3):
            ws.cell(row=row_idx, column=1, value=date)
            for col_idx, col_def in enumerate(template_columns, start=2):
                ws.cell(row=row_idx, column=col_idx, value=data_pivot.get(col_def['code'], {}).get(date, ""))
                
        output_path = self.output_folder / output_filename
        wb.save(output_path)
        logger.info(f"[OK] Output successfully saved to: {output_path}")

    def process_folders(self, folders_to_process: List[Path]):
        """
        Processes all selected folders and combines their data into a single output file.
        """
        all_excel_files = []
        for folder_path in folders_to_process:
            if folder_path.exists():
                found_files = sorted([f for f in folder_path.glob("*.xlsx") if not f.name.startswith("~$")])
                if found_files:
                    logger.info(f"--> Found {len(found_files)} files to process in '{folder_path}'")
                    all_excel_files.extend(found_files)
        
        if not all_excel_files:
            logger.error("No processable Excel files found in any selected directories.")
            return
            
        logger.info(f"\n--> Starting processing of {len(all_excel_files)} total files...")
        
        master_data_list = []
        for file_path in all_excel_files:
            data_from_one_file = self.process_single_file(file_path)
            if data_from_one_file:
                master_data_list.extend(data_from_one_file)
            
        if master_data_list:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_filename = f"JGBF_DATA_Consolidated_{timestamp}.xlsx"
            self.generate_output_file(master_data_list, output_filename)
            print(f"\n[SUCCESS] PARSING COMPLETED!")
            print(f"  • Processed {len(all_excel_files)} files from {len(folders_to_process)} folder(s).")
            print(f"  • Extracted {len(master_data_list)} total data points.")
            print(f"  > Consolidated output is available at: {self.output_folder / output_filename}")
        else:
            logger.error("Processing finished, but NO DATA could be extracted from any files.")

def main():
    print("JGBF Excel Parser - Final Production Version")
    print("=" * 55)
    current_dir = Path('.')
    exclude_dirs = {'parsed_output', '__pycache__', '.git', '.venv', '.idea'} 
    available_folders = sorted([p for p in current_dir.iterdir() if p.is_dir() and p.name not in exclude_dirs])
    
    if not available_folders:
        print("\n[ERROR] No processable subdirectories found."); sys.exit(1)
        
    folders_to_process = []
    while True:
        print("\nPlease choose which folder(s) to process:")
        for i, folder in enumerate(available_folders): print(f"  {i+1}. {folder.name}")
        print("-" * 25)
        print("  A. Process ALL listed folders")
        print("  Q. Quit")
        
        choice = input("\nEnter your choice (e.g., 1, A, Q): ").upper().strip()
        
        if choice == 'Q': print("Exiting."); sys.exit(0)
        if choice == 'A':
            folders_to_process = available_folders
            print(f"\n[OK] All {len(folders_to_process)} folders selected."); break
        try:
            choice_index = int(choice) - 1
            if 0 <= choice_index < len(available_folders):
                selected_folder = available_folders[choice_index]
                folders_to_process = [selected_folder]
                print(f"\n[OK] Folder '{selected_folder.name}' selected."); break
            else: print(f"[ERROR] Invalid number.")
        except ValueError: print("[ERROR] Invalid input.")
            
    if folders_to_process:
        parser = JGBFParser(output_folder="parsed_output")
        parser.process_folders(folders_to_process)

if __name__ == "__main__":
    main()